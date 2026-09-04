from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datos import GRUPOS, OTROS, ya_aplicado

AZUL   = "1F3864"; GRIS = "F2F2F2"; AMBAR = "FFF2CC"; VERDE = "E2EFDA"; ROJO = "FCE4EC"
F = "Arial"
wb = Workbook()

def encabezar(ws, cols):
    for i, (txt, ancho) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=txt)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

# --- Hoja 1: duplicados -------------------------------------------------------
ws = wb.active
ws.title = "MPN duplicados"
encabezar(ws, [("Grupo (MPN repetido)", 20), ("Producto", 42), ("MPN actual", 18),
               ("MPN correcto propuesto", 22), ("De donde sale la propuesta", 46),
               ("Corregido", 12), ("Notas del equipo", 30)])

fila = 2
for mpn, productos in GRUPOS:
    for j, (titulo, handle, propuesto, origen) in enumerate(productos):
        dueno = origen == "probable dueno"
        ws.cell(row=fila, column=1, value=mpn if j == 0 else "")
        ws.cell(row=fila, column=2, value=titulo)
        ws.cell(row=fila, column=3, value=mpn)
        ws.cell(row=fila, column=4, value=propuesto if not dueno else "— (dejar como esta)")
        ws.cell(row=fila, column=5, value="Es el dueno legitimo de este numero" if dueno else origen)
        hecho = (not dueno) and propuesto and ya_aplicado(handle)
        ws.cell(row=fila, column=6, value="Aplicado 24-ago" if hecho else "")
        for col in range(1, 8):
            c = ws.cell(row=fila, column=col)
            c.font = Font(name=F, size=10, bold=(col == 1))
            c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 5, 7)))
            c.border = Border(bottom=Side(style="thin", color="D9D9D9"))
            if dueno or hecho:
                c.fill = PatternFill("solid", fgColor=VERDE)
            elif propuesto:
                c.fill = PatternFill("solid", fgColor=AMBAR)
            else:
                c.fill = PatternFill("solid", fgColor=ROJO)
        fila += 1
    fila += 1  # renglon en blanco entre grupos

# --- Hoja 2: otros MPN equivocados -------------------------------------------
ws2 = wb.create_sheet("Otros MPN equivocados")
encabezar(ws2, [("Producto", 40), ("MPN actual", 18), ("MPN correcto propuesto", 22),
                ("Por que", 62), ("Corregido", 12)])
for i, (titulo, handle, actual, propuesto, porque) in enumerate(OTROS, start=2):
    ws2.cell(row=i, column=1, value=titulo)
    ws2.cell(row=i, column=2, value=actual)
    ws2.cell(row=i, column=3, value=propuesto)
    ws2.cell(row=i, column=4, value=porque)
    hecho2 = bool(propuesto) and ya_aplicado(handle)
    ws2.cell(row=i, column=5, value="Aplicado 24-ago" if hecho2 else "")
    for col in range(1, 6):
        c = ws2.cell(row=i, column=col)
        c.font = Font(name=F, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=(col in (1, 4)))
        c.border = Border(bottom=Side(style="thin", color="D9D9D9"))
        c.fill = PatternFill("solid", fgColor=VERDE if hecho2 else (AMBAR if propuesto else ROJO))

# --- Hoja 3: como usar --------------------------------------------------------
ws3 = wb.create_sheet("Como usar", 0)
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 108

# Los conteos se calculan, no se escriben a mano: la hoja no puede decir una cifra
# distinta a la que tienen las filas de al lado.
filas = [(p, o) for _, ps in GRUPOS for _, _, p, o in ps] + [(p, q) for _, _, _, p, q in OTROS]
n_hecho = sum(1 for _, ps in GRUPOS for _, h, p, o in ps if p and o != "probable dueno" and ya_aplicado(h)) \
        + sum(1 for _, h, _, p, _ in OTROS if p and ya_aplicado(h))
n_verde = sum(1 for p, o in filas if o == "probable dueno")
n_rojo  = sum(1 for p, o in filas if not p)
n_ambar = len(filas) - n_verde - n_rojo

texto = [
 ("Numeros de parte (MPN) duplicados en Shopify", "titulo"),
 (f"Detectado por specs-centinela el 22-ago-2026, sobre productos ACTIVOS. {len(GRUPOS)} grupos. "
  f"YA APLICADOS POR API el 24-ago: {n_hecho}. Quedan {n_ambar - n_hecho} por revisar y {n_rojo} sin resolver.", "sub"),
 ("", ""),
 ("Por que importa", "h"),
 ("El MPN es la llave con la que indexan Google Shopping y cualquier catalogo de contenido sindicado", "p"),
 ("(1WorldSync, Icecat). Un MPN repetido hace que le entreguen a un producto el contenido de otro:", "p"),
 ("especificaciones, fotos y todo. Mientras esto siga asi, sindicar contenido empeora el catalogo.", "p"),
 ("", ""),
 ("El patron", "h"),
 ("Al duplicar un producto en Shopify para crear otro, el numero de parte viaja con la copia y nadie lo", "p"),
 ("cambia. Es la misma raiz de las descripciones equivocadas: la Brother SP-1 y la Epson F170 comparten", "p"),
 ("MPN hasta hoy, y por eso la SP-1 tenia pegada la ficha de la F170.", "p"),
 ("", ""),
 ("Como leer la hoja", "h"),
 ("VERDE  = ya esta bien: o era el dueno legitimo del numero, o la correccion YA SE APLICO por API", "verde"),
 ("         el 24-ago (lo dice la columna Corregido). En ambos casos no hay nada que hacer.", "verde"),
 ("AMBAR  = hay propuesta de correccion, con su origen en la columna de al lado. Verificar y aplicar.", "ambar"),
 ("ROJO   = falta buscar el numero correcto en el fabricante.", "rojo"),
 ("", ""),
 ("Las dos filas que dicen CONFIRMAR", "h"),
 ("Tienen numero propuesto, pero la fuente no es del producto exacto y eso hay que cerrarlo antes de", "p"),
 ("aplicarlas. Un numero equivocado no se ve distinto a uno bueno, y por eso van marcadas:", "p"),
 ("  - Canon imageRUNNER 1643i: el folleto da 1643i = 3630C006AA y 1643iF = 3630C005AA. La 'F' es el", "p"),
 ("    fax. Hay que mirar cual de los dos equipos se vende antes de escribir el numero.", "p"),
 ("  - Caja de mantenimiento Epson T3170: el numero sale de la ficha del T3170X, no del T3170 a secas.", "p"),
 ("", ""),
 ("Donde se corrige", "h"),
 ("Shopify admin -> el producto -> Metacampos -> mm-google-shopping / mpn", "p"),
 ("Las filas que faltan se van marcando en la columna 'Corregido'; las que ya dicen 'Aplicado 24-ago'", "p"),
 ("se escribieron por API y no hay que tocarlas.", "p"),
 ("", ""),
 ("Ejemplo de fila ya resuelta", "h"),
 ("HP Smart Tank 580 | actual 4SB24A#AKY | correcto 1F3Y2A | origen: el nombre del PDF oficial de HP", "p"),
 ("que la propia tienda hospeda (MULTIFUNCIONAL-HP-SMART-TANK-580-WIRELESS-1F3Y2A.pdf)", "p"),
 ("", ""),
 ("Estado al 24-ago-2026", "h"),
 ("Se corrio specs-centinela despues de aplicar: mpn_duplicado bajo de 20 grupos a 3, y los 3 que", "p"),
 ("quedan son exactamente los que se dejaron fuera a proposito (los AMBAR/ROJO de esta hoja).", "p"),
 ("El centinela vuelve a correr solo los lunes 8:00 a.m. hora Panama.", "p"),
]
for i, (t, tipo) in enumerate(texto, start=2):
    c = ws3.cell(row=i, column=2, value=t)
    if tipo == "titulo":  c.font = Font(name=F, bold=True, size=15, color=AZUL)
    elif tipo == "sub":   c.font = Font(name=F, size=10, italic=True, color="808080")
    elif tipo == "h":     c.font = Font(name=F, bold=True, size=11, color=AZUL)
    else:                 c.font = Font(name=F, size=10)
    if tipo == "verde":   c.fill = PatternFill("solid", fgColor=VERDE)
    if tipo == "ambar":   c.fill = PatternFill("solid", fgColor=AMBAR)
    if tipo == "rojo":    c.fill = PatternFill("solid", fgColor=ROJO)

wb.save("MPN-duplicados-QSP.xlsx")
n = sum(len(p) for _, p in GRUPOS)
print(f"OK — {len(GRUPOS)} grupos, {n} productos en la hoja 1, {len(OTROS)} en la hoja 2")
